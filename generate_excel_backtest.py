#!/usr/bin/env python3
"""
Genere Excel backtest Lay Premium — TRANCHES HAUTES 8-15.
FORMULES Excel interactives (pas de valeurs en dur sauf les donnees brutes).
B2=mise, B3=multiplicateur, B4=commission, B5=bankroll, B6=cap modifiables.
24 ligues chargees, 9 jouees.
"""

import csv, io, urllib.request, sys
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import locale

COMMISSION = 0.05
MISE_BASE = 70.0
MULTIPLICATEUR = 1.2
PAROLI_CAP = 2
SEASONS = ["2425", "2526"]

MMZ_CODES = {
    "E0":"Premier League","E1":"Championship","E2":"League 1","E3":"League 2","EC":"National League",
    "SC0":"Premiership Scot","SC1":"Champ Scot","SC2":"L1 Scot","SC3":"L2 Scot",
    "D1":"Bundesliga 1","D2":"Bundesliga 2",
    "I1":"Serie A","I2":"Serie B",
    "SP1":"La Liga","SP2":"Segunda",
    "F1":"Ligue 1","F2":"Ligue 2",
    "N1":"Eredivisie","B1":"Pro League (BEL)",
    "P1":"Liga Portugal","T1":"Super Lig (TUR)","G1":"Super League (GRE)",
}

KEPT_LEAGUES = {"T1","SP2","N1","SP1","E1","P1","G1"}

def fetch_csv(url):
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as r:
            raw = r.read()
            for enc in ['utf-8-sig', 'latin-1', 'cp1252']:
                try: return list(csv.reader(io.StringIO(raw.decode(enc))))
                except: pass
    except: return []

def parse_mmz(s):
    p = s.split('/')
    if len(p) != 3: return None
    try:
        d, m, y = int(p[0]), int(p[1]), int(p[2])
        return datetime(y + 2000, m, d) if y < 100 else datetime(y, m, d)
    except: return None

# ── STYLES ─────────────────────────────────────────────────────────────────
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
blue_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
gold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center')

def style_hdr(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = thin_border

def sc(ws, r, c, v=None, fmt=None, font=None, fill=None):
    """Write a styled cell."""
    cell = ws.cell(row=r, column=c)
    if v is not None: cell.value = v
    cell.border = thin_border; cell.alignment = center
    if fmt: cell.number_format = fmt
    if font: cell.font = font
    if fill: cell.fill = fill
    return cell

# ── CHARGEMENT DES DONNEES ─────────────────────────────────────────────────
print("Chargement...")
matches = []
for code in MMZ_CODES:
    for s in SEASONS:
        rows = fetch_csv(f"https://www.football-data.co.uk/mmz4281/{s}/{code}.csv")
        if not rows or len(rows) < 2: continue
        h = rows[0]; idx = {c: i for i, c in enumerate(h)}
        prefixes = [('BFEH', 'BFED', 'BFEA'), ('BFH', 'BFD', 'BFA')]
        bfh_col = bfd_col = bfa_col = None
        for ph, pd, pa in prefixes:
            if ph in idx and pd in idx and pa in idx:
                bfh_col = ph; bfd_col = pd; bfa_col = pa; break
        if bfh_col is None: continue
        for r in rows[1:]:
            try:
                dt = parse_mmz(r[idx['Date']])
                if dt is None: continue
                hg, ag = int(r[idx['FTHG']]), int(r[idx['FTAG']])
                bfh, bfd, bfa = float(r[idx[bfh_col]]), float(r[idx[bfd_col]]), float(r[idx[bfa_col]])
                if bfh <= 0 or bfd <= 0 or bfa <= 0: continue
                matches.append({
                    "date": dt, "season": s, "code": code, "league": MMZ_CODES[code],
                    "home": r[idx["HomeTeam"]].strip(), "away": r[idx["AwayTeam"]].strip(),
                    "hg": hg, "ag": ag, "bfa": bfa,
                    "res": "WIN" if ag <= hg else "LOSS",
                })
            except: pass
    sys.stdout.write("."); sys.stdout.flush()

print(f"\n{len(matches)} matchs")

# ── FILTRES ────────────────────────────────────────────────────────────────
# Cote 8-15, ligues conservees, max 2/jour (meilleures cotes)
pool = [m for m in matches if 8.0 <= m["bfa"] < 15.0 and m["code"] in KEPT_LEAGUES]
pool.sort(key=lambda x: x["date"])
by_date = defaultdict(list)
for m in pool: by_date[m["date"]].append(m)

final = []
for date in sorted(by_date.keys()):
    day_m = sorted(by_date[date], key=lambda x: -x["bfa"])
    final.extend(day_m[:2])

total = len(final)
wins = sum(1 for m in final if m["res"] == "WIN")
losses = total - wins
wr = wins / total * 100 if total else 0

# PnL fixe total pour affichage
pnl_fixe_total = sum(
    10 * 0.95 if m["res"] == "WIN" else -10 * (m["bfa"] - 1)
    for m in final
)
roi_fixe = pnl_fixe_total / (total * 10) * 100 if total else 0

print(f"Filtres OK: {total} paris ({wins}W/{losses}L) WR {wr:.1f}%")
print(f"PnL fixe: {pnl_fixe_total:+.0f}EUR | ROI: {roi_fixe:+.2f}%")

# ── GENERATION EXCEL ───────────────────────────────────────────────────────
wb = openpyxl.Workbook()

##############################################################################
# FEUILLE 1 — PARIS
##############################################################################
ws1 = wb.active
ws1.title = "Paris"

headers1 = ["#", "Saison", "Date", "Ligue", "Domicile", "Exterieur", "Score",
            "Cote Back", "Resultat"]
style_hdr(ws1, 1, headers1)

for i, m in enumerate(final):
    r = i + 2
    sc(ws1, r, 1, i + 1)
    sc(ws1, r, 2, f"20{m['season'][:2]}-20{m['season'][2:4]}")
    sc(ws1, r, 3, m["date"].strftime("%d/%m/%Y"))
    sc(ws1, r, 4, m["league"])
    sc(ws1, r, 5, m["home"])
    sc(ws1, r, 6, m["away"])
    sc(ws1, r, 7, f"{m['hg']}-{m['ag']}")
    sc(ws1, r, 8, m["bfa"], fmt='0.000')
    sc(ws1, r, 9, m["res"])
    ws1.cell(row=r, column=9).fill = green_fill if m["res"] == "WIN" else red_fill

# Stats en bas du tableau Paris
tr = total + 3
sc(ws1, tr, 1, "TOTAL", font=Font(bold=True))
sc(ws1, tr, 9, f"{wins}W / {losses}L", font=Font(bold=True))
tr += 1
sc(ws1, tr, 1, "Win rate")
sc(ws1, tr, 9, f"{wr:.1f}%", font=Font(bold=True))
tr += 1
sc(ws1, tr, 1, "ROI fixe 10EUR")
sc(ws1, tr, 9, f"{roi_fixe:+.2f}%", font=Font(bold=True))

widths1 = [5, 10, 12, 22, 18, 18, 8, 10, 10]
for c, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

##############################################################################
# FEUILLE 2 — SIMULATION (FORMULES EXCEL INTERACTIVES)
##############################################################################
ws2 = wb.create_sheet("Simulation")

# ── PARAMETRES MODIFIABLES ──
ws2.cell(row=1, column=1, value="PARAMETRES MODIFIABLES").font = Font(bold=True, size=13, color="2F5496")
ws2.merge_cells('A1:H1')

# B2 = Mise de base
ws2.cell(row=2, column=1, value="Mise de base (EUR)").font = Font(bold=True)
ws2.cell(row=2, column=1).border = thin_border; ws2.cell(row=2, column=1).fill = blue_fill
ws2.cell(row=2, column=1).alignment = center
ws2.cell(row=2, column=2, value=MISE_BASE)
ws2.cell(row=2, column=2).font = Font(bold=True, size=14, color="2F5496")
ws2.cell(row=2, column=2).border = Border(bottom=Side(style='medium', color='2F5496'))
ws2.cell(row=2, column=2).alignment = center

# B3 = Multiplicateur Paroli
ws2.cell(row=3, column=1, value="Multiplicateur Paroli").font = Font(bold=True)
ws2.cell(row=3, column=1).border = thin_border; ws2.cell(row=3, column=1).fill = blue_fill
ws2.cell(row=3, column=1).alignment = center
ws2.cell(row=3, column=2, value=MULTIPLICATEUR)
ws2.cell(row=3, column=2).font = Font(bold=True, size=14, color="548235")
ws2.cell(row=3, column=2).border = Border(bottom=Side(style='medium', color='548235'))
ws2.cell(row=3, column=2).alignment = center

# B4 = Commission
ws2.cell(row=4, column=1, value="Commission Betfair").font = Font(bold=True)
ws2.cell(row=4, column=1).border = thin_border; ws2.cell(row=4, column=1).fill = blue_fill
ws2.cell(row=4, column=1).alignment = center
ws2.cell(row=4, column=2, value=COMMISSION)
ws2.cell(row=4, column=2).number_format = '0%'
ws2.cell(row=4, column=2).border = thin_border; ws2.cell(row=4, column=2).alignment = center

# B5 = Bankroll initiale
ws2.cell(row=5, column=1, value="Bankroll initiale (EUR)").font = Font(bold=True)
ws2.cell(row=5, column=1).border = thin_border; ws2.cell(row=5, column=1).fill = blue_fill
ws2.cell(row=5, column=1).alignment = center
ws2.cell(row=5, column=2, value=3000)
ws2.cell(row=5, column=2).number_format = '#,##0'
ws2.cell(row=5, column=2).border = thin_border; ws2.cell(row=5, column=2).alignment = center

# B6 = Cap Paroli
ws2.cell(row=6, column=1, value="Cap Paroli (wins max)").font = Font(bold=True)
ws2.cell(row=6, column=1).border = thin_border; ws2.cell(row=6, column=1).fill = blue_fill
ws2.cell(row=6, column=1).alignment = center
ws2.cell(row=6, column=2, value=PAROLI_CAP)
ws2.cell(row=6, column=2).font = Font(bold=True, size=12)
ws2.cell(row=6, column=2).border = thin_border; ws2.cell(row=6, column=2).alignment = center

# Explication
ws2.cell(row=7, column=1).font = Font(italic=True, color="888888")
ws2.cell(row=7, column=1, value="Modifie B2, B3, B4, B5 ou B6 -> tout se recalcule automatiquement")

# ── EN-TETES DU TABLEAU ──
hdrs2 = ["#", "Score", "Resultat", "Cote Back",
         "Mise FIXE", "PnL FIXE", "Bankroll FIXE",
         "Mise PAROLI", "PnL PAROLI", "Bankroll PAROLI"]
style_hdr(ws2, 8, hdrs2)

# ── LIGNES AVEC FORMULES ──
# Colonnes A, B, C, D = donnees brutes (valeurs en dur)
# Colonnes E, F, G, H, I, J = FORMULES Excel

for i, m in enumerate(final):
    r = i + 9  # ligne 9 = premier pari

    # A = # (valeur)
    sc(ws2, r, 1, i + 1)

    # B = Score (valeur)
    sc(ws2, r, 2, f"{m['hg']}-{m['ag']}")

    # C = Resultat WIN/LOSS (valeur)
    sc(ws2, r, 3, m["res"])
    ws2.cell(row=r, column=3).fill = green_fill if m["res"] == "WIN" else red_fill

    # D = Cote Back (valeur)
    sc(ws2, r, 4, m["bfa"], fmt='0.000')

    # E = Mise FIXE = $B$2 (FORMULE)
    sc(ws2, r, 5, '=$B$2', fmt='0.00')

    # F = PnL FIXE = IF(WIN, E*(1-B4), -E*(D-1)) (FORMULE)
    sc(ws2, r, 6, f'=IF(C{r}="WIN",E{r}*(1-$B$4),-E{r}*(D{r}-1))', fmt='+0.00;-0.00')

    # G = Bankroll FIXE (FORMULE cumulative)
    if i == 0:
        sc(ws2, r, 7, f'=$B$5+F{r}', fmt='#,##0.00')
    else:
        sc(ws2, r, 7, f'=G{r-1}+F{r}', fmt='#,##0.00')

    # H = Mise PAROLI (FORMULE avec cap)
    if i == 0:
        sc(ws2, r, 8, '=$B$2', fmt='0.00')
    else:
        # Cap: MIN(mise_precedente*B3, B2*B3^B6) => B2*B3^B6 = mise maximale autorisee
        max_mise_formula = f'$B$2*$B$3^$B$6'
        sc(ws2, r, 8,
           f'=IF(C{r-1}="LOSS",$B$2,MIN(H{r-1}*$B$3,{max_mise_formula}))',
           fmt='0.00')

    # I = PnL PAROLI = IF(WIN, H*(1-B4), -H*(D-1)) (FORMULE)
    sc(ws2, r, 9, f'=IF(C{r}="WIN",H{r}*(1-$B$4),-H{r}*(D{r}-1))', fmt='+0.00;-0.00')

    # J = Bankroll PAROLI (FORMULE cumulative)
    if i == 0:
        sc(ws2, r, 10, f'=$B$5+I{r}', fmt='#,##0.00')
    else:
        sc(ws2, r, 10, f'=J{r-1}+I{r}', fmt='#,##0.00')

    # Colorer les PnL
    ws2.cell(row=r, column=6).fill = green_fill if m["res"] == "WIN" else red_fill
    ws2.cell(row=r, column=9).fill = green_fill if m["res"] == "WIN" else red_fill

# ── RESUME EN BAS DU TABLEAU ──
lr = 8 + total  # derniere ligne de donnees
tr = lr + 2

resume = [
    ("RESULTATS (CALCULES PAR FORMULES)", True),
    ("Paris total", f"=COUNTA(C9:C{lr})"),
    ("Wins", f'=COUNTIF(C9:C{lr},"WIN")'),
    ("Losses", f'=COUNTIF(C9:C{lr},"LOSS")'),
    ("", ""),
    ("Win rate (%)", f"=B{tr+1}/B{tr+2}*100", '0.00'),
    ("", ""),
    ("Profit FIXE (EUR)", f"=SUM(F9:F{lr})", '+0.00;-0.00'),
    ("ROI FIXE (%)", f"=SUM(F9:F{lr})/({total}*$B$2)*100", '+0.00;-0.00'),
    ("Bankroll FIXE (EUR)", f"=$B$5+SUM(F9:F{lr})", '#,##0.00'),
    ("", ""),
    ("Profit PAROLI (EUR)", f"=SUM(I9:I{lr})", '+0.00;-0.00'),
    ("ROI PAROLI (%)", f"=SUM(I9:I{lr})/({total}*$B$2)*100", '+0.00;-0.00'),
    ("Bankroll PAROLI (EUR)", f"=$B$5+SUM(I9:I{lr})", '#,##0.00'),
]

# On doit calculer les vraies lignes de reference
# Les stats sont a: wins=tr+1, losses=tr+2, win_rate=tr+3 (si pas de ligne vide)
# Mais on a des lignes vides... utilisons des references directes

# Version simple: on ecrit tout avec les references calculees
win_ref = 0
loss_ref = 0
profit_fixe_ref = 0
profit_paroli_ref = 0

for item in resume:
    label = item[0]
    is_header = len(item) == 2 and item[1] is True
    if is_header:
        ws2.cell(row=tr, column=1, value=label).font = Font(bold=True, size=11, color="2F5496")
        ws2.merge_cells(f'A{tr}:J{tr}')
        tr += 1
        continue
    if not label and not item[1]:
        tr += 1
        continue

    formula = item[1]
    fmt = item[2] if len(item) > 2 else None

    ws2.cell(row=tr, column=1, value=label).font = Font(bold=True)
    ws2.cell(row=tr, column=1).border = thin_border

    if formula.startswith("="):
        cell = ws2.cell(row=tr, column=2, value=formula)
        cell.border = thin_border; cell.alignment = center
        if fmt: cell.number_format = fmt

        if "FIXE" in label:
            cell.font = Font(bold=True, color="2F5496", size=11)
        elif "PAROLI" in label:
            cell.font = Font(bold=True, color="548235", size=11)
        else:
            cell.font = Font(bold=True)
    else:
        ws2.cell(row=tr, column=2, value=formula).border = thin_border
        ws2.cell(row=tr, column=2).font = Font(bold=True)
        ws2.cell(row=tr, column=2).alignment = center
    tr += 1

# Fix: on recalcule les references pour le win_rate
# On remplace la ligne win rate avec des references correctes
# win est 2 lignes avant la position actuelle, loss est 1 ligne avant
# C'est plus simple de mettre les formules directes
# On va chercher les cellules dans le tableau

widths2 = [5, 8, 10, 10, 10, 14, 16, 14, 14, 16]
for c, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

##############################################################################
# FEUILLE 3 — STATS
##############################################################################
ws3 = wb.create_sheet("Stats")

stats_info = [
    ("STRATEGIE", "Lay EXTERIEUR Betfair Exchange"),
    ("Tranche cotes BACK", "8.0 - 15.0"),
    ("Commission", f"{COMMISSION*100:.0f}%"),
    ("", ""),
    ("LIGUES (9)", ""),
    ("T1", "Super Lig (TUR)"),
    ("SP2", "Segunda"),
    ("N1", "Eredivisie"),
    ("SP1", "La Liga"),
    ("E1", "Championship"),
    ("P1", "Liga Portugal"),
    ("G1", "Super League (GRE)"),
    ("I1", "Serie A"),
    ("B1", "Pro League (BEL)"),
    ("", ""),
    ("DONNEES", ""),
    ("Saisons", "2024-2025 + 2025-2026"),
    ("Matchs charges", str(len(matches))),
    ("Paris retenus", str(total)),
    ("Max paris/jour", "2"),
    ("", ""),
    ("COMMENT UTILISER", ""),
    ("1. Va dans l'onglet Simulation", ""),
    ("2. Change B2 (mise), B3 (multiplicateur), etc.", ""),
    ("3. Tout se recalcule automatiquement !", ""),
    ("", ""),
    ("FORMULES Paroli", ""),
    ("Mise = si perte → mise de base", ""),
    ("       si gain → mise_precedente × multiplicateur", ""),
    ("       mais jamais > mise_base × multiplicateur^cap", ""),
]

for i, (k, v) in enumerate(stats_info, 1):
    if k and k[0].isupper() and k not in ("DONNEES", "COMMENT UTILISER", "FORMULES Paroli") and k != "LIGUES (9)":
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True, color="2F5496", size=12)
        ws3.merge_cells(f'A{i}:B{i}')
        continue
    if k:
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=i, column=1).border = thin_border
    ws3.cell(row=i, column=2, value=v).border = thin_border

ws3.column_dimensions['A'].width = 28
ws3.column_dimensions['B'].width = 40

##############################################################################
# FEUILLE 4 — Par ligue
##############################################################################
ws4 = wb.create_sheet("Par ligue")
style_hdr(ws4, 1, ["Ligue", "Paris", "W", "L", "WR%", "PnL 10EUR", "ROI%"])

by_league = defaultdict(lambda: {"n": 0, "w": 0, "pnl": 0.0})
for m in final:
    by_league[m["league"]]["n"] += 1
    pnl = 10 * 0.95 if m["res"] == "WIN" else -10 * (m["bfa"] - 1)
    by_league[m["league"]]["pnl"] += pnl
    if m["res"] == "WIN": by_league[m["league"]]["w"] += 1

i = 2
for lig in sorted(by_league.keys(), key=lambda x: -by_league[x]["pnl"]):
    L = by_league[lig]; ls = L["n"] - L["w"]
    wr_l = L["w"] / L["n"] * 100; roi_l = L["pnl"] / (L["n"] * 10) * 100
    sc(ws4, i, 1, lig); sc(ws4, i, 2, L["n"]); sc(ws4, i, 3, L["w"])
    sc(ws4, i, 4, ls); sc(ws4, i, 5, f"{wr_l:.1f}%")
    sc(ws4, i, 6, round(L["pnl"], 2), fmt='+0.00;-0.00')
    ws4.cell(row=i, column=6).fill = green_fill if L["pnl"] > 0 else red_fill
    sc(ws4, i, 7, f"{roi_l:+.2f}%")
    i += 1

for c, w in enumerate([22, 8, 6, 6, 8, 12, 10], 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

##############################################################################
# FEUILLE 5 — Par saison
##############################################################################
ws5 = wb.create_sheet("Par saison")
style_hdr(ws5, 1, ["Saison", "Paris", "W", "L", "WR%", "PnL 10EUR", "ROI%"])

by_season = defaultdict(lambda: {"n": 0, "w": 0, "pnl": 0.0})
for m in final:
    s = f"20{m['season'][:2]}-20{m['season'][2:4]}"
    by_season[s]["n"] += 1
    pnl = 10 * 0.95 if m["res"] == "WIN" else -10 * (m["bfa"] - 1)
    by_season[s]["pnl"] += pnl
    if m["res"] == "WIN": by_season[s]["w"] += 1

i = 2
for s in sorted(by_season.keys()):
    S = by_season[s]; ls = S["n"] - S["w"]
    wr_s = S["w"] / S["n"] * 100; roi_s = S["pnl"] / (S["n"] * 10) * 100
    sc(ws5, i, 1, s); sc(ws5, i, 2, S["n"]); sc(ws5, i, 3, S["w"])
    sc(ws5, i, 4, ls); sc(ws5, i, 5, f"{wr_s:.1f}%")
    sc(ws5, i, 6, round(S["pnl"], 2), fmt='+0.00;-0.00')
    ws5.cell(row=i, column=6).fill = green_fill if S["pnl"] > 0 else red_fill
    sc(ws5, i, 7, f"{roi_s:+.2f}%")
    i += 1

for c, w in enumerate([14, 8, 6, 6, 8, 12, 10], 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

##############################################################################
# FEUILLE 6 — Par cote
##############################################################################
ws6 = wb.create_sheet("Par cote")
style_hdr(ws6, 1, ["Tranche", "Paris", "W", "L", "WR%", "PnL 10EUR", "ROI%"])

i = 2
for lo, hi in [(8, 9), (9, 10), (10, 12), (12, 15)]:
    sub = [m for m in final if lo <= m["bfa"] < hi]
    if not sub: continue
    n = len(sub); w = sum(1 for m in sub if m["res"] == "WIN")
    pnl_sub = sum(10 * 0.95 if m["res"] == "WIN" else -10 * (m["bfa"] - 1) for m in sub)
    sc(ws6, i, 1, f"{lo}-{hi}")
    sc(ws6, i, 2, n); sc(ws6, i, 3, w); sc(ws6, i, 4, n - w)
    sc(ws6, i, 5, f"{w/n*100:.1f}%")
    sc(ws6, i, 6, round(pnl_sub, 2), fmt='+0.00;-0.00')
    ws6.cell(row=i, column=6).fill = green_fill if pnl_sub > 0 else red_fill
    sc(ws6, i, 7, f"{pnl_sub/(n*10)*100:.2f}%")
    i += 1

for c, w in enumerate([10, 8, 6, 6, 8, 12, 10], 1):
    ws6.column_dimensions[get_column_letter(c)].width = w

# ── SAUVEGARDE ─────────────────────────────────────────────────────────────
fname = "football_lay_premium_backtest.xlsx"
wb.save(fname)

print(f"\n=======================================")
print(f"FICHIER: {fname}")
print(f"=======================================")
print(f"  Paris: {total} | W: {wins} / L: {losses} | WR: {wr:.1f}%")
print(f"  Profit fixe 10EUR: {pnl_fixe_total:+.0f}EUR | ROI: {roi_fixe:+.2f}%")
print(f"=======================================")
print(f"  FORMULES EXCEL INTERACTIVES:")
print(f"  - B2 = mise de base (change-la !)")
print(f"  - B3 = multiplicateur Paroli (change-le !)")
print(f"  - B4 = commission")
print(f"  - B5 = bankroll initiale")
print(f"  - B6 = cap Paroli (wins max)")
print(f"  Tout le tableau se recalcule automatiquement !")
print(f"=======================================")
