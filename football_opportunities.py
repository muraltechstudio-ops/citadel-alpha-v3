#!/usr/bin/env python3
"""
OPPORTUNITY FINDER — Lay Exterieur Betfair
Recherche les matchs du jour / à venir selon nos critères stricts.

Critères de sélection :
1. Marché : Lay EXTERIEUR (parier que l'équipe extérieure ne gagne PAS)
2. Cotes BACK extérieur (BF_A) entre 8.0 et 15.0
3. 9 ligues seulement : Super Lig TUR, Segunda, Eredivisie, La Liga,
   Championship, Liga Portugal, Super League GRE, Serie A, Pro League BEL
4. Max 2 paris/jour (meilleures cotes)
5. Bankroll : 70€ base, x1.2 Paroli, cap 2 wins, capital 3000€
"""

import csv, io, urllib.request, sys, os, json
from datetime import datetime, date, timedelta
from collections import defaultdict
import openpyxl
import urllib.parse

# ── CONFIG ─────────────────────────────────────────────────────────────────
COMMISSION = 0.05
MISE_BASE = 70.0
MULTIPLICATEUR = 1.2
CAP_PAROLI = 2
BANKROLL_INIT = 3000.0
COTE_MIN = 8.0
COTE_MAX = 15.0

CHEMIN_ETAT = "paroli_state.json"

# Les 7 ligues jouees (Pro League BEL: -14.4%, Serie A: +2.1% -> trop juste)
LIGUES_ACTIVES = {
    "T1": "Super Lig (TUR)",
    "SP2": "Segunda",
    "N1": "Eredivisie",
    "SP1": "La Liga",
    "E1": "Championship",
    "P1": "Liga Portugal",
    "G1": "Super League (GRE)",
}

# Toutes les ligues pour chargement (les actives filtrees apres)
ALL_CODES = {
    "E0": "Premier League", "E1": "Championship", "E2": "League 1",
    "E3": "League 2", "EC": "National League",
    "SC0": "Premiership Scot", "SC1": "Champ Scot",
    "SC2": "L1 Scot", "SC3": "L2 Scot",
    "D1": "Bundesliga 1", "D2": "Bundesliga 2",
    "I1": "Serie A", "I2": "Serie B",
    "SP1": "La Liga", "SP2": "Segunda",
    "F1": "Ligue 1", "F2": "Ligue 2",
    "N1": "Eredivisie", "B1": "Pro League (BEL)",
    "P1": "Liga Portugal", "T1": "Super Lig (TUR)",
    "G1": "Super League (GRE)",
}

# Tournois internationaux / Coupes (format Excel)
# La Coupe du Monde 2026 est dispo en xlsx chez football-data.co.uk
URL_WC_XLSX = "https://www.football-data.co.uk/WorldCup2026.xlsx"

SEASONS = ["2526"]

# TheOddsAPI
ODDS_API_KEY = "53a174078d66229028e9bf942c6868f9"

# Mapping codes ligues -> TheOddsAPI sport keys
# football-data.co.uk code -> sport_key TheOdds
LIGUES_ODDS_API = {
    "T1": "soccer_turkey_super_lig",
    "SP2": None,
    "N1": "soccer_netherlands_eredivisie",
    "SP1": "soccer_spain_la_liga",
    "E1": "soccer_england_championship",
    "P1": "soccer_portugal_primeira_liga",
    "G1": None,
}

# ── FONCTIONS ──────────────────────────────────────────────────────────────
def fetch_csv(url):
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as f:
            raw = f.read()
        for enc in ['utf-8-sig', 'latin-1', 'cp1252']:
            try: return list(csv.reader(io.StringIO(raw.decode(enc))))
            except: pass
    except:
        pass
    return []

def parse_mmz(s):
    """Parse date format DD/MM/YYYY."""
    p = s.split('/')
    if len(p) != 3:
        return None
    try:
        d, m, y = int(p[0]), int(p[1]), int(p[2])
        return datetime(y + 2000, m, d) if y < 100 else datetime(y, m, d)
    except:
        return None

def determiner_saison():
    """Determine la saison en cours a partir de la date."""
    maintenant = date.today()
    if maintenant.month >= 8:
        return f"{maintenant.year % 100:02d}{(maintenant.year + 1) % 100:02d}"
    else:
        return f"{(maintenant.year - 1) % 100:02d}{maintenant.year % 100:02d}"

def charger_etat_paroli():
    """Charge l'etat Paroli depuis le fichier JSON."""
    if os.path.exists(CHEMIN_ETAT):
        try:
            with open(CHEMIN_ETAT, 'r') as f:
                return json.load(f)
        except:
            pass
    return {"streak": 0, "bankroll": BANKROLL_INIT, "mise_courante": MISE_BASE, "historique": []}

def sauver_etat_paroli(etat):
    """Sauve l'etat Paroli dans le fichier JSON."""
    with open(CHEMIN_ETAT, 'w') as f:
        json.dump(etat, f, indent=2)

def calculer_mise_paroli(etat):
    """Calcule la mise Paroli en fonction du streak et du cap."""
    mult = MULTIPLICATEUR ** min(etat["streak"], CAP_PAROLI)
    mise = round(MISE_BASE * mult, 2)
    return mise

def enregistrer_resultat(etat, resultat, cote, mise):
    """Enregistre le resultat d'un pari et met a jour l'etat Paroli."""
    if resultat == "WIN":
        pnl = round(mise * (1 - COMMISSION), 2)
        etat["streak"] += 1
    else:
        pnl = round(-mise * (cote - 1), 2)
        etat["streak"] = 0

    etat["bankroll"] = round(etat["bankroll"] + pnl, 2)
    etat["mise_courante"] = calculer_mise_paroli(etat)
    return pnl

def match_est_disponible(r, idx):
    """Verifie si un match a des cotes Betfair et est a venir (pas de resultat)."""
    # Verifier presence cotes Betfair
    prefixes = [('BFEH', 'BFED', 'BFEA'), ('BFH', 'BFD', 'BFA')]
    for ph, pd, pa in prefixes:
        if ph in idx and pd in idx and pa in idx:
            break
    else:
        return None

    try:
        bfa = float(r[idx.get(ph, -1)]) if ph in idx else None
        if bfa is None or bfa <= 0:
            return None

        # Verifier si le match a deja ete joue
        # Si FTHG ou FTAG sont vides, le match n'a pas encore eu lieu
        fthg = r[idx.get("FTHG", -1)].strip() if "FTHG" in idx else ""
        ftag = r[idx.get("FTAG", -1)].strip() if "FTAG" in idx else ""

        deja_joue = False
        if fthg and ftag:
            try:
                int(fthg)
                int(ftag)
                deja_joue = True
            except:
                deja_joue = False

        return {
            "code": ph[2] if len(ph) >= 3 else "",
            "bfa": bfa,
            "deja_joue": deja_joue,
            "fthg": fthg,
            "ftag": ftag,
        }
    except:
        return None


def detecter_saison_suivante(season):
    """Retourne la saison suivante dans le format mmz."""
    y1 = int(season[:2])
    y2 = int(season[2:4])
    if y2 >= 99:
        return f"{y2:02d}{(y2+1)%100:02d}"
    return f"{y2:02d}{(y2+1)%100:02d}"


def charger_wc_xlsx():
    """Charge les matchs de la Coupe du Monde 2026 depuis le fichier Excel."""
    from openpyxl import load_workbook
    try:
        req = urllib.request.Request(URL_WC_XLSX, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=60) as f:
            raw = f.read()
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        h = {}
        for c in range(1, ws.max_column + 1):
            h[str(ws.cell(1, c).value).strip().lower()] = c

        wc_matches = []
        for r in range(2, ws.max_row + 1):
            try:
                date_val = ws.cell(r, h.get('date', 4)).value
                if not date_val: continue
                if isinstance(date_val, datetime):
                    dt = date_val
                else:
                    dt = datetime.strptime(str(date_val)[:10], '%Y-%m-%d')

                home = str(ws.cell(r, h.get('home', 2)).value or '').strip()
                away = str(ws.cell(r, h.get('away', 3)).value or '').strip()
                if not home or not away: continue

                bfeh_c = h.get('betfair_exch-h', 34)
                bfed_c = h.get('betfair_exch-d', 35)
                bfea_c = h.get('betfair_exch-a', 36)

                bfh = float(ws.cell(r, bfeh_c).value or 0)
                bfd = float(ws.cell(r, bfed_c).value or 0)
                bfa = float(ws.cell(r, bfea_c).value or 0)
                if bfh <= 0 or bfd <= 0 or bfa <= 0: continue

                hg_v = ws.cell(r, h.get('hgft', 6)).value
                ag_v = ws.cell(r, h.get('agft', 7)).value
                hg_str = str(hg_v or '').strip()
                ag_str = str(ag_v or '').strip()
                score_renseigne = hg_str.isdigit() and ag_str.isdigit() and hg_str and ag_str
                # Un match est a venir si le score EST renseigne ET la date est dans le futur
                a_venir = not score_renseigne and dt > datetime.now()
                hg = int(hg_str) if hg_str.isdigit() else 0
                ag = int(ag_str) if ag_str.isdigit() else 0

                wc_matches.append({
                    'date': dt, 'code': 'WC', 'league': 'Coupe du Monde 2026',
                    'home': home, 'away': away,
                    'bfa': bfa, 'bfh': bfh, 'bfd': bfd,
                    'a_venir': a_venir,
                    'fthg': hg_str if hg_str else '?',
                    'ftag': ag_str if ag_str else '?',
                    'season': '2526',
                })
                if len(wc_matches) >= 500: break
            except:
                pass
        return wc_matches
    except Exception as e:
        print(f'\n  Note: Coupe du Monde non chargee ({e})')
        return []


def charger_odds_api():
    """Charge les matchs a venir depuis TheOddsAPI pour nos ligues actives."""
    matchs = []
    for code, sport_key in LIGUES_ODDS_API.items():
        if sport_key is None:
            continue
        url = f"https://api.the-odds-api.com/v4/sports/{sport_key}/odds/?apiKey={ODDS_API_KEY}&regions=eu,uk&markets=h2h"
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=30) as f:
                raw = f.read()
            data = json.loads(raw)
            if not isinstance(data, list):
                continue
            for m in data:
                try:
                    dt = datetime.fromisoformat(m['commence_time'].replace('Z', '+00:00'))
                    home = m['home_team']
                    away = m['away_team']
                    bfa = None
                    bfh = None
                    bfd = None
                    for bm in m.get('bookmakers', []):
                        if bm['key'] in ('betfair', 'pinnacle'):
                            for o in bm['markets'][0]['outcomes']:
                                if o['name'] == away:
                                    bfa = o['price']
                                elif o['name'] == home:
                                    bfh = o['price']
                                elif o['name'] == 'Draw':
                                    bfd = o['price']
                    if bfa and bfh and bfd and 8.0 <= bfa < 15.0:
                        matchs.append({
                            'date': dt, 'code': code, 'league': LIGUES_ACTIVES[code],
                            'home': home, 'away': away,
                            'bfa': bfa, 'bfh': bfh, 'bfd': bfd,
                            'a_venir': True,
                            'fthg': '?', 'ftag': '?',
                            'season': '2526',
                            'source': 'odds_api',
                        })
                except:
                    pass
            sys.stdout.write(f'[{sport_key[:20]}] ')
            sys.stdout.flush()
        except Exception as e:
            pass
    return matchs


# ── MAIN ───────────────────────────────────────────────────────────────────
def main():
    # Parsing argument
    args = sys.argv[1:]
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except:
        pass

    etat = charger_etat_paroli()

    # Gestion des commandes
    if args and args[0] == "--status":
        mise = calculer_mise_paroli(etat)
        print(f"Etat Paroli: {etat['streak']}W consec")
        print(f"Mise courante: {mise:.1f}EUR")
        print(f"Bankroll: {etat['bankroll']:.0f}EUR")
        return

    if args and args[0] == "--reset":
        etat = {"streak": 0, "bankroll": BANKROLL_INIT, "mise_courante": MISE_BASE, "historique": []}
        sauver_etat_paroli(etat)
        print("Etat Paroli reinitialise.")
        return

    if args and args[0] == "--result":
        if len(args) < 3:
            print("Usage: --result WIN|LOSS cote")
            print("Exemple: --result WIN 12.5")
            return
        resultat = args[1].upper()
        try:
            cote = float(args[2])
        except:
            print("Cote invalide.")
            return
        if resultat not in ("WIN", "LOSS"):
            print("Resultat doit etre WIN ou LOSS.")
            return
        mise = calculer_mise_paroli(etat)
        pnl = enregistrer_resultat(etat, resultat, cote, mise)
        etat["historique"].append({"date": date.today().isoformat(), "resultat": resultat, "cote": cote, "mise": mise, "pnl": pnl})
        etat["bankroll"] = round(etat["bankroll"], 2)
        sauver_etat_paroli(etat)
        print(f"Pari enregistre: {resultat} (cote {cote})")
        print(f"  Mise: {mise:.1f}EUR | PnL: {pnl:+.1f}EUR")
        print(f"  Bankroll: {etat['bankroll']:.1f}EUR | Streak: {etat['streak']}W consec")
        print(f"  Prochaine mise: {calculer_mise_paroli(etat):.1f}EUR")
        return

    # Forcer saison via -s
    saison = determiner_saison()
    saisons_pour_scan = [saison]
    for i, a in enumerate(args):
        if a == "-s" and i + 1 < len(args):
            saisons_pour_scan = [args[i + 1]]
            saison = args[i + 1]
            break

    aujourd_hui = date.today()
    print("=" * 65)
    print(f"  OPPORTUNITY FINDER — Lay Exterieur")
    print(f"  {aujourd_hui.strftime('%d/%m/%Y')} | Saison 20{saison[:2]}-20{saison[2:4]}")
    print("=" * 65)

    # ── CHARGEMENT ETAT PAROLI ──
    etat = charger_etat_paroli()
    mise_proposee = calculer_mise_paroli(etat)
    print(f"\n  Etat Paroli: {etat['streak']}W consec | "
          f"Bankroll: {etat['bankroll']:.0f}EUR | "
          f"Mise courante: {mise_proposee:.1f}EUR")

    # ── CHARGEMENT DONNEES ──
    # Ajouter la saison suivante si on est fin de saison (mai-juillet)
    if aujourd_hui.month >= 5 and aujourd_hui.month <= 7:
        if len(saisons_pour_scan) == 1:
            saisons_pour_scan.append(detecter_saison_suivante(saison))

    matchs_trouves = []
    ligues_chargees = 0

    for ssn in saisons_pour_scan:
        for code, nom_ligue in ALL_CODES.items():
            url = f"https://www.football-data.co.uk/mmz4281/{ssn}/{code}.csv"
            rows = fetch_csv(url)
            if not rows or len(rows) < 2:
                continue

            h = rows[0]
            idx = {c: i for i, c in enumerate(h)}

            # Verifier si cette ligue a les cotes Betfair
            prefixes = [('BFEH', 'BFED', 'BFEA'), ('BFH', 'BFD', 'BFA')]
            bfh_col = bfd_col = bfa_col = None
            for ph, pd, pa in prefixes:
                if ph in idx and pd in idx and pa in idx:
                    bfh_col, bfd_col, bfa_col = ph, pd, pa
                    break
            if bfh_col is None:
                continue

            ligues_chargees += 1
            for r in rows[1:]:
                try:
                    dt = parse_mmz(r[idx['Date']])
                    if dt is None:
                        continue
                    match_date = dt.date()

                    bfa = float(r[idx[bfa_col]])
                    if bfa <= 0:
                        continue

                    # Verifier si c'est un match a venir (pas de resultat)
                    fthg_str = r[idx.get("FTHG", -1)].strip() if "FTHG" in idx else ""
                    ftag_str = r[idx.get("FTAG", -1)].strip() if "FTAG" in idx else ""
                    a_venir = True
                    try:
                        if fthg_str and ftag_str:
                            int(fthg_str); int(ftag_str)
                            a_venir = False
                    except:
                        pass

                    home = r[idx["HomeTeam"]].strip()
                    away = r[idx["AwayTeam"]].strip()

                    matchs_trouves.append({
                        "date": dt,
                        "code": code,
                        "league": ALL_CODES[code],
                        "home": home,
                        "away": away,
                        "bfa": bfa,
                        "bfh": float(r[idx[bfh_col]]),
                        "bfd": float(r[idx[bfd_col]]),
                        "a_venir": a_venir,
                        "fthg": fthg_str if fthg_str else "?",
                        "ftag": ftag_str if ftag_str else "?",
                        "season": ssn,
                    })
                except:
                    pass
            sys.stdout.write(".")
            sys.stdout.flush()

    # Charger aussi la Coupe du Monde 2026
    wc_matchs = charger_wc_xlsx()
    if wc_matchs:
        matchs_trouves.extend(wc_matchs)
        print(f"\n  Coupe du Monde 2026: {len(wc_matchs)} matchs charges")

    # Charger les matchs a venir depuis TheOddsAPI
    print("\n  TheOddsAPI: scan des matchs a venir...")
    odds_matchs = charger_odds_api()
    if odds_matchs:
        matchs_trouves.extend(odds_matchs)
        print(f"\n  TheOddsAPI: {len(odds_matchs)} matchs a venir ajoutes")
    else:
        print("  TheOddsAPI: aucun match a venir trouve")

    print(f"\n  {ligues_chargees} fichiers charges, {len(matchs_trouves)} matchs trouves")

    # ── VERIFICATION DES SAISONS ──
    # Afficher le calendrier des saisons trouvees
    saisons_trouvees = set(m["season"] for m in matchs_trouves)
    print(f"  Saisons trouvees: {', '.join(sorted([f'20{s[:2]}-20{s[2:4]}' for s in saisons_trouvees]))}")

    # ── FILTRES STRICTS ──
    print(f"\n  --- APPLICATION DES FILTRES ---")

    # Filtre 1: ligues actives uniquement
    actifs = [m for m in matchs_trouves if m["code"] in LIGUES_ACTIVES]
    print(f"  Ligues actives (9): {len(actifs)} matchs")

    # Filtre 2: cote 8.0 - 15.0
    bonne_cote = [m for m in actifs if COTE_MIN <= m["bfa"] < COTE_MAX]
    print(f"  Cote [{COTE_MIN}-{COTE_MAX}[ : {len(bonne_cote)} matchs")

    # Parmi les matchs eligibles, separer ceux a venir et ceux deja joues
    a_venir = [m for m in bonne_cote if m["a_venir"]]
    deja_joues = [m for m in bonne_cote if not m["a_venir"]]

    print(f"  Matchs a venir (oportunites): {len(a_venir)}")
    print(f"  Matchs deja joues (historique): {len(deja_joues)}")

    # ── SELECTION MAX 2/JOUR ──
    if a_venir:
        # Grouper par date
        par_date = defaultdict(list)
        for m in a_venir:
            par_date[m["date"].date()].append(m)

        selections = []
        for d in sorted(par_date.keys()):
            jour_m = sorted(par_date[d], key=lambda x: -x["bfa"])
            selections.extend(jour_m[:2])

        print(f"\n  {'=' * 65}")
        print(f"  SELECTIONS DU JOUR")
        print(f"  {'=' * 65}")

        aujourd_hui_sel = [m for m in selections if m["date"].date() == aujourd_hui]
        futures = [m for m in selections if m["date"].date() > aujourd_hui]

        if aujourd_hui_sel:
            print(f"\n  AUJOURD'HUI ({aujourd_hui.strftime('%d/%m/%Y')}):")
            print(f"  {'Ligue':<22} {'Domicile':<20} {'Exterieur':<20} {'Cote':>6} {'H/D/A':>8}")
            print(f"  {'-' * 78}")
            for m in aujourd_hui_sel:
                print(f"  {m['league']:<22} {m['home']:<20} {m['away']:<20} {m['bfa']:>6.3f} {m['bfh']:>5.2f}/{m['bfd']:>3.2f}/{m['bfa']:>3.2f}")

            # Instructions de jeu
            print(f"\n  INSTRUCTIONS DE JEU:")
            for m in aujourd_hui_sel:
                print(f"  🎯 {m['home']} vs {m['away']} ({m['league']})")
                print(f"     LAY {m['away']} (parier que {m['away']} ne gagne PAS)")
                print(f"     Cote LAY ≈ {m['bfa']:.3f} (Back price)")
                print(f"     Mise recommandee: {mise_proposee:.1f}EUR (Paroli)")
            print(f"\n     📊 Paroli: {etat['streak']}W consec | "
                  f"Mise: {mise_proposee:.1f}EUR | "
                  f"Bankroll dispo: {etat['bankroll']:.0f}EUR")
            print(f"     ⚠️  Perte max si LOSS: ~{mise_proposee * (max(m['bfa'] for m in aujourd_hui_sel) - 1):.0f}EUR")
        else:
            print(f"\n  ❌ Aucun match eligible aujourd'hui.")

        if futures:
            print(f"\n  PROCHAINEMENT:")
            for d in sorted(set(m["date"].date() for m in futures)):
                jour_m = [m for m in futures if m["date"].date() == d]
                print(f"  {d.strftime('%d/%m/%Y')}:")
                for m in jour_m:
                    print(f"    {m['league']:<22} {m['home']:<20} vs {m['away']:<20} "
                          f"Cote {m['bfa']:.3f}")
    else:
        print(f"\n  ❌ AUCUN MATCH A VENIR avec nos criteres.")
        print(f"  Cause possible: hors saison (juin-juillet) ou赛季 pas encore commence.")

        # Afficher quand meme les matchs joues avec nos criteres pour info
        if deja_joues:
            # Grouper par date
            par_date = defaultdict(list)
            for m in deja_joues[:50]:
                par_date[m["date"].date()].append(m)

            print(f"\n  Derniers matchs joues avec nos criteres (echantillon):")
            for d in sorted(par_date.keys(), reverse=True)[:5]:
                jour_m = sorted(par_date[d], key=lambda x: -x["bfa"])[:2]
                for m in jour_m:
                    perf = "✅" if int(m["fthg"]) >= int(m["ftag"]) else "❌"
                    print(f"  {d.strftime('%d/%m/%Y')} | {perf} {m['league']:<18} "
                          f"{m['home']:<18} {m['away']:<18} "
                          f"Cote {m['bfa']:.3f} | Score {m['fthg']}-{m['ftag']}")

        # Estimation de la prochaine saison
        prochaine_saison = detecter_saison_suivante(saison)
        mois = 8 if aujourd_hui.month < 8 else aujourd_hui.month
        annee = aujourd_hui.year if aujourd_hui.month < 8 else aujourd_hui.year + 1
        print(f"\n  Prochaine saison: 20{prochaine_saison[:2]}-20{prochaine_saison[2:4]}")
        print(f"  Premieres journees: ~aout {annee}")
        print(f"  Utilise: -s 2627 pour forcer une saison")

    # ── RECAP CONFIG ──
    print(f"\n  {'=' * 65}")
    print(f"  CONFIG ACTUELLE")
    print(f"  {'=' * 65}")
    print(f"  Strategie:         Lay EXTERIEUR (outsider ne gagne pas)")
    print(f"  Tranche cotes:     {COTE_MIN}-{COTE_MAX}")
    print(f"  Ligues:            {len(LIGUES_ACTIVES)} (Super Lig, Segunda, Eredivisie, La Liga,")
    print(f"                     Championship, Liga Portugal, Super League GRE,")
    print(f"                     Serie A, Pro League BEL)")
    print(f"  Paris/jour:        Max 2 (meilleures cotes)")
    print(f"  Mise base:         {MISE_BASE:.0f}EUR")
    print(f"  Paroli:            x{MULTIPLICATEUR}, cap {CAP_PAROLI} wins")
    print(f"  Commission:        {COMMISSION*100:.0f}%")
    print(f"  Bankroll init:     {BANKROLL_INIT:.0f}EUR")
    print(f"  Etat Paroli:       {etat['streak']}W consec")
    print(f"  Bankroll dispo:    {etat['bankroll']:.0f}EUR")

    # ── COMMANDES ──
    print(f"\n  {'=' * 65}")
    print(f"  COMMANDES")
    print(f"  {'=' * 65}")
    print(f"  python football_opportunities.py                    # Scan defaut")
    print(f"  python football_opportunities.py -s 2627            # Forcer saison")
    print(f"  python football_opportunities.py --result WIN 12.5  # Enregistrer WIN + cote")
    print(f"  python football_opportunities.py --result LOSS 9.4  # Enregistrer LOSS + cote")
    print(f"  python football_opportunities.py --status           # Voir etat Paroli")
    print(f"  python football_opportunities.py --reset            # Reset etat Paroli")
    print()


if __name__ == "__main__":
    main()
