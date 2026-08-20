#!/usr/bin/env python3
"""
Interface Web — Lay Exterieur Betfair
Dashboard HTML pour scanner, suivre le Paroli et enregistrer les resultats.
"""

import csv, io, urllib.request, sys, os, json
from datetime import datetime, date, timedelta
from collections import defaultdict
import openpyxl
from flask import Flask, render_template_string, jsonify, request

COMMISSION = 0.05
MISE_BASE = 70.0
MULTIPLICATEUR = 1.2
CAP_PAROLI = 2
BANKROLL_INIT = 3000.0
COTE_MIN = 8.0
COTE_MAX = 15.0
ODDS_API_KEY = "53a174078d66229028e9bf942c6868f9"
CHEMIN_ETAT = "paroli_state.json"
URL_WC_XLSX = "https://www.football-data.co.uk/WorldCup2026.xlsx"

MMZ_CODES = {
    "E0": "Premier League", "E1": "Championship", "E2": "League 1",
    "E3": "League 2", "EC": "National League",
    "SC0": "Premiership Scot", "SC1": "Champ Scot",
    "SC2": "L1 Scot", "SC3": "L2 Scot",
    "D1": "Bundesliga 1", "D2": "Bundesliga 2",
    "I1": "Serie A", "I2": "Serie B",
    "SP1": "La Liga", "SP2": "Segunda",
    "F1": "Ligue 1", "F2": "Ligue 2",
    "N1": "Eredivisie", "B1": "Pro League (BEL)",
    "P1": "Liga Portugal", "T1": "Super Lig (TUR)",
    "G1": "Super League (GRE)",
}

LIGUES_ACTIVES = {
    "T1": "Super Lig (TUR)",
    "SP2": "Segunda",
    "N1": "Eredivisie",
    "SP1": "La Liga",
    "E1": "Championship",
    "P1": "Liga Portugal",
    "G1": "Super League (GRE)",
}

LIGUES_ODDS_API = {
    "T1": "soccer_turkey_super_lig",
    "N1": "soccer_netherlands_eredivisie",
    "SP1": "soccer_spain_la_liga",
    "E1": "soccer_england_championship",
    "P1": "soccer_portugal_primeira_liga",
    "I1": "soccer_italy_serie_a",
}

SEASONS = ["2526"]

# ── FONCTIONS METIER ───────────────────────────────────────────────────

def fetch_csv(url):
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as f:
            raw = f.read()
        for enc in ['utf-8-sig', 'latin-1', 'cp1252']:
            try: return list(csv.reader(io.StringIO(raw.decode(enc))))
            except: pass
    except: return []

def parse_mmz(s):
    p = s.split('/')
    if len(p) != 3: return None
    try:
        d, m, y = int(p[0]), int(p[1]), int(p[2])
        return datetime(y + 2000, m, d) if y < 100 else datetime(y, m, d)
    except: return None

def charger_etat():
    if os.path.exists(CHEMIN_ETAT):
        try:
            with open(CHEMIN_ETAT) as f:
                return json.load(f)
        except: pass
    return {"streak": 0, "bankroll": BANKROLL_INIT, "mise_courante": MISE_BASE, "historique": []}

def sauver_etat(etat):
    with open(CHEMIN_ETAT, 'w') as f:
        json.dump(etat, f, indent=2)

def calculer_mise(etat):
    return round(MISE_BASE * (MULTIPLICATEUR ** min(etat["streak"], CAP_PAROLI)), 2)

def calculer_pnl(resultat, mise, cote):
    if resultat == "WIN":
        return round(mise * (1 - COMMISSION), 2)
    else:
        return round(-mise * (cote - 1), 2)

def charger_wc():
    try:
        req = urllib.request.Request(URL_WC_XLSX, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=60) as f:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()))
        ws = wb.active
        h = {}
        for c in range(1, ws.max_column + 1):
            h[str(ws.cell(1, c).value).strip().lower()] = c
        matchs = []
        for r in range(2, ws.max_row + 1):
            try:
                d = ws.cell(r, h.get('date', 4)).value
                dt = d if isinstance(d, datetime) else datetime.strptime(str(d)[:10], '%Y-%m-%d')
                home = str(ws.cell(r, h.get('home', 2)).value or '').strip()
                away = str(ws.cell(r, h.get('away', 3)).value or '').strip()
                if not home or not away: continue
                bfa = float(ws.cell(r, h.get('betfair_exch-a', 36)).value or 0)
                bfh = float(ws.cell(r, h.get('betfair_exch-h', 34)).value or 0)
                bfd = float(ws.cell(r, h.get('betfair_exch-d', 35)).value or 0)
                if bfa <= 0 or bfh <= 0 or bfd <= 0: continue
                hg = str(ws.cell(r, h.get('hgft', 6)).value or '').strip()
                ag = str(ws.cell(r, h.get('agft', 7)).value or '').strip()
                a_venir = not (hg.isdigit() and ag.isdigit()) and dt > datetime.now()
                matchs.append({
                    'date': dt, 'code': 'WC', 'league': 'Coupe du Monde 2026',
                    'home': home, 'away': away, 'bfa': bfa, 'bfh': bfh, 'bfd': bfd,
                    'a_venir': a_venir, 'fthg': hg if hg else '?', 'ftag': ag if ag else '?',
                })
            except: pass
        return matchs
    except: return []

def charger_odds():
    matchs = []
    for code, sport_key in LIGUES_ODDS_API.items():
        if sport_key is None or code not in LIGUES_ACTIVES: continue
        url = f"https://api.the-odds-api.com/v4/sports/{sport_key}/odds/?apiKey={ODDS_API_KEY}&regions=eu,uk&markets=h2h&bookmakers=pinnacle,bet365"
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=30) as f:
                data = json.loads(f.read())
            if not isinstance(data, list): continue
            for m in data:
                try:
                    dt = datetime.fromisoformat(m['commence_time'].replace('Z', '+00:00'))
                    home, away = m['home_team'], m['away_team']
                    bfa = bfh = bfd = None
                    for bm in m.get('bookmakers', []):
                        for o in bm['markets'][0]['outcomes']:
                            if o['name'] == away: bfa = o['price']
                            elif o['name'] == home: bfh = o['price']
                            elif o['name'] == 'Draw': bfd = o['price']
                    if bfa and bfh and bfd:
                        matchs.append({
                            'date': dt, 'code': code, 'league': LIGUES_ACTIVES.get(code, code),
                            'home': home, 'away': away, 'bfa': bfa, 'bfh': bfh, 'bfd': bfd,
                            'a_venir': True, 'fthg': '?', 'ftag': '?',
                        })
                except: pass
        except: pass
    return matchs

def charger_donnees():
    """Charge et filtre tous les matchs."""
    aujourd_hui = date.today()
    saison = f"{aujourd_hui.year % 100:02d}{(aujourd_hui.year + 1) % 100:02d}" if aujourd_hui.month >= 8 else f"{(aujourd_hui.year - 1) % 100:02d}{aujourd_hui.year % 100:02d}"

    matchs = []

    # MMZ
    for code in MMZ_CODES:
        rows = fetch_csv(f"https://www.football-data.co.uk/mmz4281/{saison}/{code}.csv")
        if not rows or len(rows) < 2: continue
        h, idx = rows[0], {c: i for i, c in enumerate(rows[0])}
        prefixes = [('BFEH','BFED','BFEA'),('BFH','BFD','BFA')]
        bfh_c = bfd_c = bfa_c = None
        for ph, pd, pa in prefixes:
            if ph in idx and pd in idx and pa in idx: bfh_c, bfd_c, bfa_c = ph, pd, pa; break
        if bfh_c is None: continue
        for r in rows[1:]:
            try:
                dt = parse_mmz(r[idx['Date']])
                if dt is None: continue
                bfa = float(r[idx[bfa_c]]); bfh = float(r[idx[bfh_c]]); bfd = float(r[idx[bfd_c]])
                if bfa <= 0 or bfh <= 0 or bfd <= 0: continue
                fthg = r[idx.get("FTHG", -1)].strip() if "FTHG" in idx else ""
                ftag = r[idx.get("FTAG", -1)].strip() if "FTAG" in idx else ""
                a_venir = True
                try:
                    if fthg and ftag: int(fthg); int(ftag); a_venir = False
                except: pass
                matchs.append({
                    'date': dt, 'code': code, 'league': MMZ_CODES[code],
                    'home': r[idx["HomeTeam"]].strip(), 'away': r[idx["AwayTeam"]].strip(),
                    'bfa': bfa, 'bfh': bfh, 'bfd': bfd,
                    'a_venir': a_venir, 'fthg': fthg if fthg else '?', 'ftag': ftag if ftag else '?',
                })
            except: pass

    # WC
    for m in charger_wc():
        matchs.append(m)

    # Odds API
    for m in charger_odds():
        matchs.append(m)

    return matchs

def filtrer_opportunites(matchs):
    """Filtre les matchs pour nos criteres et retourne les selections du jour."""
    # Filtrer ligues actives + cote 8-15
    actifs = [m for m in matchs if m['code'] in LIGUES_ACTIVES and COTE_MIN <= m['bfa'] < COTE_MAX]

    # Filtrer a venir seulement
    a_venir = [m for m in actifs if m['a_venir']]

    # Max 2/jour (meilleures cotes)
    by_date = defaultdict(list)
    for m in a_venir:
        by_date[m['date'].date()].append(m)

    selections = []
    for d in sorted(by_date.keys()):
        day_m = sorted(by_date[d], key=lambda x: -x['bfa'])
        selections.extend(day_m[:2])

    return selections

def get_stats_historique(etat):
    """Calcule les stats a partir de l'historique."""
    h = etat.get('historique', [])
    total = len(h)
    wins = sum(1 for r in h if r['resultat'] == 'WIN')
    losses = total - wins
    pnl = sum(r.get('pnl', 0) for r in h)
    return {'total': total, 'wins': wins, 'losses': losses, 'pnl': round(pnl, 2)}


# ── FLASK APP ────────────────────────────────────────────────────────

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Lay Exterieur — Dashboard</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
               background: #0f1923; color: #e0e6ed; padding: 20px; }
        .container { max-width: 1200px; margin: 0 auto; }

        .header {
            background: linear-gradient(135deg, #1a2a3a, #0d1b2a);
            border: 1px solid #2a3a4a;
            border-radius: 16px; padding: 24px 32px; margin-bottom: 24px;
            display: flex; justify-content: space-between; align-items: center;
        }
        .header h1 { font-size: 24px; font-weight: 700; background: linear-gradient(90deg, #4fc3f7, #00e676);
                     -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
        .header .date { color: #8899aa; font-size: 14px; }

        .cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 24px; }
        .card {
            background: #1a2a3a; border: 1px solid #2a3a4a; border-radius: 12px; padding: 20px;
            text-align: center; transition: 0.2s;
        }
        .card:hover { border-color: #4fc3f7; transform: translateY(-2px); }
        .card .label { font-size: 12px; color: #8899aa; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px; }
        .card .value { font-size: 28px; font-weight: 700; }
        .card .value.green { color: #00e676; }
        .card .value.red { color: #ff5252; }
        .card .value.blue { color: #4fc3f7; }
        .card .value.gold { color: #ffd740; }
        .card .sub { font-size: 12px; color: #667788; margin-top: 4px; }

        .section { background: #1a2a3a; border: 1px solid #2a3a4a; border-radius: 12px; padding: 20px; margin-bottom: 24px; }
        .section h2 { font-size: 18px; font-weight: 600; margin-bottom: 16px; color: #4fc3f7; }

        table { width: 100%; border-collapse: collapse; font-size: 14px; }
        th { text-align: left; padding: 10px 12px; color: #8899aa; font-size: 12px; text-transform: uppercase;
             letter-spacing: 0.5px; border-bottom: 1px solid #2a3a4a; }
        td { padding: 10px 12px; border-bottom: 1px solid #1e2e3e; }
        tr:hover td { background: rgba(79, 195, 247, 0.05); }
        .tag { display: inline-block; padding: 2px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; }
        .tag.win { background: rgba(0, 230, 118, 0.15); color: #00e676; }
        .tag.loss { background: rgba(255, 82, 82, 0.15); color: #ff5252; }
        .cote-badge { background: #0f1923; padding: 2px 8px; border-radius: 6px; font-family: monospace; font-size: 13px; }
        .league-badge { background: #2a3a4a; padding: 2px 8px; border-radius: 6px; font-size: 11px; color: #aabbcc; }

        .btn {
            display: inline-block; padding: 6px 16px; border-radius: 8px; border: none; cursor: pointer;
            font-size: 13px; font-weight: 600; transition: 0.2s; text-decoration: none;
        }
        .btn-win { background: rgba(0, 230, 118, 0.15); color: #00e676; border: 1px solid #00e676; }
        .btn-win:hover { background: #00e676; color: #0f1923; }
        .btn-loss { background: rgba(255, 82, 82, 0.15); color: #ff5252; border: 1px solid #ff5252; }
        .btn-loss:hover { background: #ff5252; color: #fff; }
        .btn-refresh { background: rgba(79, 195, 247, 0.15); color: #4fc3f7; border: 1px solid #4fc3f7; }
        .btn-refresh:hover { background: #4fc3f7; color: #0f1923; }
        .btn-reset { background: rgba(255, 82, 82, 0.1); color: #ff5252; border: 1px solid #ff5252; font-size: 11px; padding: 4px 12px; }

        .empty { text-align: center; padding: 40px; color: #667788; }
        .empty .big { font-size: 48px; margin-bottom: 12px; }
        .empty p { font-size: 14px; }

        .config-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; }
        .config-item { text-align: center; padding: 12px; background: #0f1923; border-radius: 8px; }
        .config-item .label { font-size: 11px; color: #8899aa; }
        .config-item .value { font-size: 16px; font-weight: 600; margin-top: 4px; }

        .historique { max-height: 300px; overflow-y: auto; }
        .historique::-webkit-scrollbar { width: 6px; }
        .historique::-webkit-scrollbar-track { background: #0f1923; border-radius: 3px; }
        .historique::-webkit-scrollbar-thumb { background: #2a3a4a; border-radius: 3px; }

        .notification {
            position: fixed; top: 20px; right: 20px; padding: 12px 24px; border-radius: 10px;
            font-weight: 600; z-index: 1000; animation: slideIn 0.3s ease;
        }
        .notification.success { background: #1b5e20; color: #a5d6a7; border: 1px solid #2e7d32; }
        .notification.error { background: #b71c1c; color: #ef9a9a; border: 1px solid #c62828; }
        @keyframes slideIn { from { transform: translateX(100px); opacity: 0; } to { transform: translateX(0); opacity: 1; } }

        @media (max-width: 600px) {
            .header { flex-direction: column; gap: 8px; text-align: center; }
            .cards { grid-template-columns: repeat(2, 1fr); }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div>
                <h1>Lay Exterieur</h1>
                <div class="date">{{ date_aujourdhui }}</div>
            </div>
            <div>
                <a href="/" class="btn btn-refresh">⟳ Raffraichir</a>
                <a href="/reset" class="btn btn-reset" onclick="return confirm('Reset du Paroli ?')">Reset Paroli</a>
            </div>
        </div>

        <!-- CARTES STATS -->
        <div class="cards">
            <div class="card">
                <div class="label">Bankroll</div>
                <div class="value gold">{{ "%.0f"|format(etat.bankroll) }}€</div>
                <div class="sub">depart 3000€</div>
            </div>
            <div class="card">
                <div class="label">Streak</div>
                <div class="value blue">{{ etat.streak }}W</div>
                <div class="sub">consecutives</div>
            </div>
            <div class="card">
                <div class="label">Prochaine mise</div>
                <div class="value blue">{{ "%.1f"|format(mise_courante) }}€</div>
                <div class="sub">base {{ mise_base }}€ ×{{ multiplicateur }}{% if etat.streak > 0 %}^{% if etat.streak > cap_paroli %}{{ cap_paroli }}{% else %}{{ etat.streak }}{% endif %}{% endif %}</div>
            </div>
            <div class="card">
                <div class="label">Profit total</div>
                <div class="value {% if stats_histo.pnl >= 0 %}green{% else %}red{% endif %}">
                    {{ "%+.0f"|format(stats_histo.pnl) }}€
                </div>
                <div class="sub">{{ stats_histo.wins }}W / {{ stats_histo.losses }}L</div>
            </div>
        </div>

        <!-- OPPORTUNITES DU JOUR -->
        <div class="section">
            <h2>🎯 Opportunités du jour</h2>
            {% if opportunites %}
            <table>
                <tr><th>Ligue</th><th>Domicile</th><th>Exterieur</th><th>Cote Back</th><th>Action</th></tr>
                {% for m in opportunites %}
                <tr>
                    <td><span class="league-badge">{{ m.league[:18] }}</span></td>
                    <td>{{ m.home }}</td>
                    <td><strong>{{ m.away }}</strong></td>
                    <td><span class="cote-badge">{{ "%.3f"|format(m.bfa) }}</span></td>
                    <td>
                        <a href="/result/WIN/{{ m.bfa }}" class="btn btn-win" onclick="return confirm('WIN: {{ m.away }} n\\'a pas gagne ?')">Lay gagne</a>
                        <a href="/result/LOSS/{{ m.bfa }}" class="btn btn-loss" onclick="return confirm('LOSS: {{ m.away }} a gagne ?')">Lay perdu</a>
                    </td>
                </tr>
                {% endfor %}
            </table>
            {% else %}
            <div class="empty">
                <div class="big">🏖️</div>
                <p>Aucune opportunité aujourd'hui — hors saison (juin-juillet).<br>Reviens en août quand les championnats reprennent !</p>
            </div>
            {% endif %}
        </div>

        <!-- DERNIERS RESULTATS -->
        <div class="section">
            <h2>📜 Derniers résultats</h2>
            {% if etat.historique %}
            <div class="historique">
                <table>
                    <tr><th>Date</th><th>Resultat</th><th>Cote</th><th>Mise</th><th>PnL</th></tr>
                    {% for r in etat.historique[-20:]|reverse %}
                    <tr>
                        <td>{{ r.date[:10] }}</td>
                        <td><span class="tag {{ 'win' if r.resultat=='WIN' else 'loss' }}">{{ r.resultat }}</span></td>
                        <td>{{ "%.2f"|format(r.cote) }}</td>
                        <td>{{ "%.1f"|format(r.mise) }}€</td>
                        <td style="color: {{ '#00e676' if r.pnl >= 0 else '#ff5252' }}">{{ "%+.0f"|format(r.pnl) }}€</td>
                    </tr>
                    {% endfor %}
                </table>
            </div>
            {% else %}
            <div class="empty">
                <p>Aucun résultat enregistré. <br>Utilise les boutons au-dessus après chaque match.</p>
            </div>
            {% endif %}
        </div>

        <!-- CONFIG -->
        <div class="section">
            <h2>⚙️ Configuration</h2>
            <div class="config-grid">
                <div class="config-item"><div class="label">Stratégie</div><div class="value" style="font-size:13px">Lay EXTERIEUR</div></div>
                <div class="config-item"><div class="label">Tranche cotes</div><div class="value" style="color:#4fc3f7">{{ cote_min }}-{{ cote_max }}</div></div>
                <div class="config-item"><div class="label">Ligues</div><div class="value" style="font-size:13px">{{ nb_ligues }}</div></div>
                <div class="config-item"><div class="label">Mise base</div><div class="value" style="color:#ffd740">{{ mise_base }}€</div></div>
                <div class="config-item"><div class="label">Multiplicateur</div><div class="value" style="color:#4fc3f7">×{{ multiplicateur }}</div></div>
                <div class="config-item"><div class="label">Cap wins</div><div class="value" style="color:#00e676">{{ cap_paroli }}</div></div>
                <div class="config-item"><div class="label">Commission</div><div class="value" style="font-size:13px">{{ commission_pct }}%</div></div>
                <div class="config-item"><div class="label">Paris/jour</div><div class="value" style="font-size:13px">Max 2</div></div>
            </div>
        </div>
    </div>
</body>
</html>
"""

# ── FLASK ROUTES ─────────────────────────────────────────────────────

app = Flask(__name__)

@app.route('/')
def index():
    etat = charger_etat()
    mise_courante = calculer_mise(etat)
    stats_histo = get_stats_historique(etat)

    matchs = charger_donnees()
    opportunites = filtrer_opportunites(matchs)

    # Filtrer pour aujourd'hui seulement
    aujourd_hui = date.today()
    opportunites = [m for m in opportunites if m['date'].date() == aujourd_hui]

    return render_template_string(HTML_TEMPLATE,
        etat=etat,
        mise_courante=mise_courante,
        mise_base=MISE_BASE,
        multiplicateur=MULTIPLICATEUR,
        cap_paroli=CAP_PAROLI,
        cote_min=COTE_MIN,
        cote_max=COTE_MAX,
        commission_pct=int(COMMISSION * 100),
        nb_ligues=len(LIGUES_ACTIVES),
        stats_histo=stats_histo,
        opportunites=opportunites,
        date_aujourdhui=aujourd_hui.strftime('%d/%m/%Y'),
    )

@app.route('/result/<resultat>/<float:cote>')
def enregistrer(resultat, cote):
    if resultat not in ('WIN', 'LOSS'):
        return jsonify({'error': 'WIN ou LOSS attendu'}), 400
    etat = charger_etat()
    mise = calculer_mise(etat)
    pnl = calculer_pnl(resultat, mise, cote)

    if resultat == 'WIN':
        etat['streak'] += 1
    else:
        etat['streak'] = 0

    etat['bankroll'] = round(etat['bankroll'] + pnl, 2)
    etat['mise_courante'] = calculer_mise(etat)
    etat.setdefault('historique', []).append({
        'date': datetime.now().isoformat(),
        'resultat': resultat,
        'cote': cote,
        'mise': mise,
        'pnl': pnl,
    })
    sauver_etat(etat)
    return jsonify({
        'status': 'ok',
        'resultat': resultat,
        'mise': mise,
        'pnl': pnl,
        'bankroll': etat['bankroll'],
        'streak': etat['streak'],
        'prochaine_mise': calculer_mise(etat),
    })

@app.route('/reset')
def reset():
    etat = {"streak": 0, "bankroll": BANKROLL_INIT, "mise_courante": MISE_BASE, "historique": []}
    sauver_etat(etat)
    return '<script>window.location.href="/";</script>'

@app.route('/api/status')
def api_status():
    etat = charger_etat()
    return jsonify({
        'streak': etat['streak'],
        'bankroll': etat['bankroll'],
        'mise_courante': calculer_mise(etat),
        'historique_count': len(etat.get('historique', [])),
    })


if __name__ == '__main__':
    print("=" * 55)
    print("  LAY EXTERIEUR — Dashboard")
    print("  http://127.0.0.1:5000")
    print("=" * 55)
    print("  Ctrl+C pour quitter")
    print("=" * 55)
    app.run(host='127.0.0.1', port=5000, debug=False)
